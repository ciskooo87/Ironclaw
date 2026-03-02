import csv
from pathlib import Path
from typing import Any, Dict, List, Tuple

from .utils import norm_key


def map_value(normalized: Dict[str, Any], aliases: Dict[str, List[str]], field: str, default: Any = "") -> Any:
    for c in aliases.get(field, [field]):
        if c in normalized and str(normalized[c]).strip() != "":
            return normalized[c]
    return default


def normalize_row(row: Dict[str, Any], source_file: str, idx: int, aliases: Dict[str, List[str]]) -> Dict[str, Any]:
    normalized = {norm_key(k): v for k, v in row.items()}
    return {
        "periodo": map_value(normalized, aliases, "periodo", ""),
        "unidade": map_value(normalized, aliases, "unidade", ""),
        "kpi": map_value(normalized, aliases, "kpi", ""),
        "valor_atual": map_value(normalized, aliases, "valor_atual", "0"),
        "meta": map_value(normalized, aliases, "meta", "0"),
        "variacao": map_value(normalized, aliases, "variacao", ""),
        "observacao": map_value(normalized, aliases, "observacao", ""),
        "fonte_arquivo": source_file,
        "linha": idx,
    }


def normalize_headers(headers: List[str]) -> List[str]:
    return [norm_key(str(h or "")) for h in headers]


def validate_required_headers(headers: List[str], aliases: Dict[str, List[str]], required_fields: List[str]) -> List[str]:
    header_set = set(headers)
    missing = []
    for field in required_fields:
        if not set(aliases.get(field, [field])).intersection(header_set):
            missing.append(field)
    return missing


def load_csv(file_path: Path, aliases: Dict[str, List[str]], required_fields: List[str]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    rows, issues = [], []
    with file_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = normalize_headers(reader.fieldnames or [])
        missing_headers = validate_required_headers(headers, aliases, required_fields)
        if missing_headers:
            issues.append({"type": "missing_required_headers", "missing_headers": missing_headers, "fonte_arquivo": file_path.name})
            return rows, issues
        for idx, row in enumerate(reader, start=2):
            rows.append(normalize_row(row, file_path.name, idx, aliases))
    return rows, issues


def _canonical_from_sheet_row(sheet_name: str, row: Dict[str, Any], source_file: str, idx: int) -> List[Dict[str, Any]]:
    s = norm_key(sheet_name)
    periodo = row.get("mes_/_ano") or row.get("dia/mes_/_ano") or row.get("mes/ano") or ""

    def rec(kpi: str, valor: Any, meta: Any = 0, unidade: str | None = None, variacao: Any = "", obs: Any = "") -> Dict[str, Any]:
        return {
            "periodo": periodo,
            "unidade": unidade or sheet_name,
            "kpi": kpi,
            "valor_atual": valor if valor not in (None, "") else 0,
            "meta": meta if meta not in (None, "") else 0,
            "variacao": variacao if variacao is not None else "",
            "observacao": obs if obs is not None else "",
            "fonte_arquivo": f"{source_file}::{sheet_name}",
            "linha": idx,
        }

    out: List[Dict[str, Any]] = []

    if s == "faturamento":
        out.append(rec("faturamento_bruto", row.get("faturamento_bruto"), meta=row.get("faturamento_previsto")))
    elif s == "faturamento_diario":
        out.append(rec("faturamento_bruto_diario", row.get("faturamento_bruto"), meta=row.get("faturamento_previsto")))
    elif s == "faturamento_x_qde_vendida":
        out.append(rec("faturamento_bruto", row.get("faturamento_bruto_r$"), unidade=row.get("grupo_de_linha") or sheet_name))
        out.append(rec("margem_liquida_percent", row.get("margem_liquida_%"), unidade=row.get("grupo_de_linha") or sheet_name))
    elif s == "curva_abc_clientes":
        out.append(rec("vendas_cliente_valor", row.get("valor"), unidade=row.get("clientes") or sheet_name))
        out.append(rec("margem_percentual", row.get("margem_%"), unidade=row.get("clientes") or sheet_name))
    elif s == "participacao_por_estado":
        out.append(rec("faturamento_por_uf", row.get("faturamento"), unidade=row.get("uf") or sheet_name, variacao=row.get("part_s/_fat_%")))
    elif s == "devolucoes":
        out.append(rec("devolucoes_valor", row.get("devolucoes"), meta=row.get("faturamento"), variacao=row.get("%_dev_s/fat")))
    elif s == "devolucoes_motivos":
        out.append(rec("devolucao_por_motivo", row.get("valor"), unidade=row.get("motivos") or sheet_name, obs=row.get("cliente")))
    elif s == "estoques":
        out.append(rec("estoque_custo_total", row.get("custo_total"), unidade=row.get("departamento") or sheet_name, obs=row.get("material")))
    elif s == "fopag":
        out.append(rec("fopag_total_geral", row.get("total_geral"), unidade=row.get("departamento") or sheet_name, obs=row.get("cargo")))
    elif s == "contas_a_pagar":
        for mes in ["janeiro", "fevereiro", "marco", "abril", "maio", "junho", "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]:
            if row.get(mes) not in (None, ""):
                out.append(rec(f"contas_pagar_{mes}", row.get(mes), unidade=row.get("natureza") or sheet_name, obs=row.get("classificacao")))
    elif s == "contas_a_receber":
        out.append(rec("contas_receber_valor", row.get("valor"), unidade=row.get("instituicao") or sheet_name, obs=row.get("modalidade")))
    elif s == "despesas_financeiras":
        out.append(rec("despesa_financeira", row.get("valor"), unidade=sheet_name, obs=row.get("despesas")))
    elif s == "endividamento":
        out.append(rec("endividamento_a_vencer", row.get("a_vencer"), unidade=row.get("instituicao") or sheet_name, obs=row.get("modalidade")))
        out.append(rec("endividamento_vencido", row.get("vencido"), unidade=row.get("instituicao") or sheet_name, obs=row.get("modalidade")))
    elif s == "dre_i":
        out.append(rec("resultado_liquido_exercicio", row.get("(=)_resultado_liquido_do_exercicio"), unidade=sheet_name))
        out.append(rec("receita_operacional_bruta", row.get("(+)_receita_operacional_bruta"), unidade=sheet_name))

    return [r for r in out if str(r.get("valor_atual", "")).strip() != ""]


def load_xlsx(file_path: Path, aliases: Dict[str, List[str]], required_fields: List[str]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    issues: List[Dict[str, Any]] = []
    try:
        from openpyxl import load_workbook
    except Exception:
        issues.append({"type": "missing_dependency", "dependency": "openpyxl", "fonte_arquivo": file_path.name})
        return [], issues

    wb = load_workbook(file_path, read_only=True, data_only=True)
    all_rows: List[Dict[str, Any]] = []

    for ws in wb.worksheets:
        data = list(ws.iter_rows(values_only=True))
        if not data:
            continue
        headers = normalize_headers([str(h or "") for h in data[0]])

        # Path A: already canonical sheet
        missing_headers = validate_required_headers(headers, aliases, required_fields)
        if not missing_headers:
            for i, values in enumerate(data[1:], start=2):
                row_raw = {headers[c]: values[c] for c in range(min(len(headers), len(values)))}
                all_rows.append(normalize_row(row_raw, f"{file_path.name}::{ws.title}", i, aliases))
            continue

        # Path B: known business sheet -> convert to canonical records
        for i, values in enumerate(data[1:], start=2):
            row_raw = {headers[c]: values[c] for c in range(min(len(headers), len(values)))}
            canonical_rows = _canonical_from_sheet_row(ws.title, row_raw, file_path.name, i)
            all_rows.extend(canonical_rows)

    if not all_rows:
        issues.append({"type": "no_usable_rows", "fonte_arquivo": file_path.name, "message": "Nenhuma linha utilizável encontrada no XLSX."})

    return all_rows, issues
