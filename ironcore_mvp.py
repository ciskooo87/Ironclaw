#!/usr/bin/env python3
import argparse
import csv
import datetime as dt
import glob
import json
import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

ROOT = Path(__file__).resolve().parent
DEFAULT_SOURCES = ROOT / "sources"
DEFAULT_PROCESSED = ROOT / "processed"
DEFAULT_OUTPUTS = ROOT / "outputs"
DEFAULT_CONFIG = ROOT / "config"
DEFAULT_LOGS = ROOT / "logs"

DEFAULT_REQUIRED_FIELDS = ["periodo", "unidade", "kpi", "valor_atual", "meta"]


def setup_logging(log_dir: Path, run_id: str | None = None) -> Path:
    ts = run_id or dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    log_path = log_dir / f"run-{ts}.log"
    log_dir.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[logging.FileHandler(log_path), logging.StreamHandler()],
    )
    return log_path


def to_float(val: Any) -> float:
    if val is None:
        return 0.0
    s = str(val).strip()
    if not s:
        return 0.0
    s = s.replace("R$", "").replace("%", "").strip()
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def norm_key(key: str) -> str:
    k = key.strip().lower()
    k = re.sub(r"\s+", "_", k)
    return (
        k.replace("á", "a").replace("à", "a").replace("â", "a").replace("ã", "a")
        .replace("é", "e").replace("ê", "e").replace("í", "i")
        .replace("ó", "o").replace("ô", "o").replace("õ", "o")
        .replace("ú", "u").replace("ç", "c")
    )


def load_mappings(config_dir: Path) -> Tuple[Dict[str, List[str]], List[str]]:
    path = config_dir / "mappings.json"
    aliases: Dict[str, List[str]] = {
        "periodo": ["periodo", "mes", "data"],
        "unidade": ["unidade", "area", "bu"],
        "kpi": ["kpi", "indicador", "metric"],
        "valor_atual": ["valor_atual", "valor", "atual"],
        "meta": ["meta", "target"],
        "variacao": ["variacao", "delta"],
        "observacao": ["observacao", "nota", "comentario"],
    }
    required_fields = DEFAULT_REQUIRED_FIELDS.copy()

    if path.exists():
        try:
            content = json.loads(path.read_text(encoding="utf-8"))
            file_aliases = content.get("aliases", {})
            aliases = {
                canonical: [norm_key(canonical)] + [norm_key(a) for a in arr]
                for canonical, arr in file_aliases.items()
            }
            required_fields = [norm_key(x) for x in content.get("required_fields", DEFAULT_REQUIRED_FIELDS)]
        except Exception as e:
            logging.warning("Falha ao carregar mappings.json (%s). Usando padrões.", e)
    return aliases, required_fields


def map_value(normalized: Dict[str, Any], aliases: Dict[str, List[str]], field: str, default: Any = "") -> Any:
    for c in aliases.get(field, [field]):
        if c in normalized and str(normalized[c]).strip() != "":
            return normalized[c]
    return default


def normalize_row(row: Dict[str, Any], source_file: str, idx: int, aliases: Dict[str, List[str]]) -> Dict[str, Any]:
    normalized = {norm_key(k): v for k, v in row.items()}
    return {
        "periodo": map_value(normalized, aliases, "periodo", ""),
        "unidade": map_value(normalized, aliases, "unidade", ""),
        "kpi": map_value(normalized, aliases, "kpi", ""),
        "valor_atual": map_value(normalized, aliases, "valor_atual", "0"),
        "meta": map_value(normalized, aliases, "meta", "0"),
        "variacao": map_value(normalized, aliases, "variacao", ""),
        "observacao": map_value(normalized, aliases, "observacao", ""),
        "fonte_arquivo": source_file,
        "linha": idx,
    }


def normalize_headers(headers: List[str]) -> List[str]:
    return [norm_key(str(h or "")) for h in headers]


def validate_required_headers(headers: List[str], aliases: Dict[str, List[str]], required_fields: List[str]) -> List[str]:
    header_set = set(headers)
    missing = []
    for field in required_fields:
        if not set(aliases.get(field, [field])).intersection(header_set):
            missing.append(field)
    return missing


def load_csv(file_path: Path, aliases: Dict[str, List[str]], required_fields: List[str]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    rows, issues = [], []
    with file_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = normalize_headers(reader.fieldnames or [])
        missing_headers = validate_required_headers(headers, aliases, required_fields)
        if missing_headers:
            issues.append({"type": "missing_required_headers", "missing_headers": missing_headers, "fonte_arquivo": file_path.name, "message": f"Arquivo sem colunas obrigatórias: {', '.join(missing_headers)}"})
            return rows, issues
        for idx, row in enumerate(reader, start=2):
            rows.append(normalize_row(row, file_path.name, idx, aliases))
    return rows, issues


def load_xlsx(file_path: Path, aliases: Dict[str, List[str]], required_fields: List[str]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    issues: List[Dict[str, Any]] = []
    try:
        from openpyxl import load_workbook
    except Exception:
        issues.append({"type": "missing_dependency", "dependency": "openpyxl", "fonte_arquivo": file_path.name, "message": "openpyxl não encontrado para leitura de XLSX"})
        return [], issues

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return [], issues

    headers = normalize_headers([str(h or "") for h in data[0]])
    missing_headers = validate_required_headers(headers, aliases, required_fields)
    if missing_headers:
        issues.append({"type": "missing_required_headers", "missing_headers": missing_headers, "fonte_arquivo": file_path.name, "message": f"Arquivo sem colunas obrigatórias: {', '.join(missing_headers)}"})
        return [], issues

    rows: List[Dict[str, Any]] = []
    for i, values in enumerate(data[1:], start=2):
        row_raw = {headers[c]: values[c] for c in range(min(len(headers), len(values)))}
        rows.append(normalize_row(row_raw, file_path.name, i, aliases))
    return rows, issues


def validate_rows(rows: List[Dict[str, Any]], required_fields: List[str]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    valid, issues = [], []
    for row in rows:
        missing = [f for f in required_fields if str(row.get(f, "")).strip() == ""]
        if missing:
            issues.append({"type": "missing_required_fields", "missing": missing, "fonte_arquivo": row.get("fonte_arquivo"), "linha": row.get("linha"), "message": f"Linha sem campos obrigatórios: {', '.join(missing)}"})
            continue
        valid.append(row)
    return valid, issues


def load_rules(config_dir: Path) -> List[Dict[str, Any]]:
    yaml_path = config_dir / "risk_rules.yaml"
    if yaml_path.exists():
        try:
            import yaml  # type: ignore
            content = yaml.safe_load(yaml_path.read_text(encoding="utf-8")) or {}
            return content.get("rules", [])
        except Exception as e:
            logging.warning("Falha ao carregar risk_rules.yaml (%s). Usando padrão.", e)

    return [
        {"name": "desvio_critico", "condition": "valor_atual < meta * 0.8", "impact": 5, "urgency": 5, "description": "Desvio crítico vs meta"},
        {"name": "abaixo_meta", "condition": "valor_atual < meta", "impact": 4, "urgency": 4, "description": "KPI abaixo da meta"},
    ]


def eval_condition(cond: str, ctx: Dict[str, Any]) -> bool:
    allowed = {"valor_atual": ctx["valor_atual"], "meta": ctx["meta"], "variacao": to_float(ctx.get("variacao", 0))}
    try:
        return bool(eval(cond, {"__builtins__": {}}, allowed))
    except Exception:
        return False


def level(score: int) -> str:
    if score >= 16:
        return "Crítico"
    if score >= 9:
        return "Alto"
    if score >= 4:
        return "Médio"
    return "Baixo"


def build_facts(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    facts = []
    for r in rows:
        valor = to_float(r["valor_atual"])
        meta = to_float(r["meta"])
        facts.append({
            "periodo": r["periodo"], "unidade": r["unidade"], "kpi": r["kpi"],
            "valor_atual": valor, "meta": meta, "desvio": valor - meta,
            "variacao": r.get("variacao", ""), "observacao": r.get("observacao", ""),
            "evidence": {"source_file": r["fonte_arquivo"], "line": r["linha"]},
        })
    return facts


def build_risks(facts: List[Dict[str, Any]], rules: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    for f in facts:
        key = f"{f['periodo']}|{f['unidade']}|{f['kpi']}"
        for rule in rules:
            if not eval_condition(rule.get("condition", "False"), f):
                continue
            impact = int(rule.get("impact", 3))
            urgency = int(rule.get("urgency", 3))
            score = impact * urgency
            if key not in grouped:
                grouped[key] = {
                    "periodo": f["periodo"], "unidade": f["unidade"], "kpi": f["kpi"],
                    "valor_atual": f["valor_atual"], "meta": f["meta"],
                    "impact": impact, "urgency": urgency, "score": score, "level": level(score),
                    "triggered_rules": [rule.get("name")],
                    "triggered_descriptions": [rule.get("description", "Risco identificado")],
                    "action_owner_suggestion": f"Responsável de {f['unidade']}",
                    "action_5w2h": {
                        "what": f"Recuperar KPI {f['kpi']} para meta",
                        "why": rule.get("description", "Mitigar risco operacional/financeiro"),
                        "where": f["unidade"],
                        "when": "Próximo ciclo semanal",
                        "who": f"Gestor(a) de {f['unidade']}",
                        "how": "Plano de ação focado em causa raiz + cadência semanal",
                        "how_much": "A definir",
                    },
                    "evidence": f["evidence"],
                }
            else:
                g = grouped[key]
                g["triggered_rules"].append(rule.get("name"))
                g["triggered_descriptions"].append(rule.get("description", "Risco identificado"))
                if score > g["score"]:
                    g["impact"], g["urgency"], g["score"], g["level"] = impact, urgency, score, level(score)
                    g["action_5w2h"]["why"] = rule.get("description", "Mitigar risco operacional/financeiro")

    risks = list(grouped.values())
    for r in risks:
        r["triggered_rules"] = sorted(list(set(r["triggered_rules"])))
        r["triggered_descriptions"] = sorted(list(set(r["triggered_descriptions"])))
    risks.sort(key=lambda x: x["score"], reverse=True)
    return risks


def render_markdown(summary: Dict[str, Any], top_risks: List[Dict[str, Any]]) -> str:
    lines = [
        "# Comitê de Turnaround — Saída IRONCORE (MVP)", "", "## Resumo executivo",
        f"- Registros processados: **{summary['processed']}**",
        f"- Riscos únicos identificados: **{summary['risks']}**",
        f"- Críticos/Altos: **{summary['critical_high']}**",
        f"- Issues de dados: **{summary['issues']}**", "", "## Top riscos priorizados",
    ]
    for i, r in enumerate(top_risks, start=1):
        w = r["action_5w2h"]
        lines.extend([
            f"### {i}. [{r['level']}] {r['kpi']} — {r['unidade']}",
            f"- Score: {r['score']} (impacto {r['impact']} x urgência {r['urgency']})",
            f"- Situação: valor atual {r['valor_atual']} vs meta {r['meta']}",
            f"- Regras acionadas: {', '.join(r['triggered_rules'])}",
            f"- Motivos: {'; '.join(r['triggered_descriptions'])}",
            "- Plano 5W2H:",
            f"  - What: {w['what']}", f"  - Why: {w['why']}", f"  - Where: {w['where']}",
            f"  - When: {w['when']}", f"  - Who: {w['who']}", f"  - How: {w['how']}",
            f"  - How much: {w['how_much']}",
            f"- Evidência: {r['evidence']['source_file']}#L{r['evidence']['line']}", "",
        ])
    return "\n".join(lines)


def ensure_dirs(*dirs: Path) -> None:
    for d in dirs:
        d.mkdir(parents=True, exist_ok=True)


def run(
    max_risks: int = 10,
    input_dir: Path = DEFAULT_SOURCES,
    processed_dir: Path = DEFAULT_PROCESSED,
    output_dir: Path = DEFAULT_OUTPUTS,
    config_dir: Path = DEFAULT_CONFIG,
    log_dir: Path = DEFAULT_LOGS,
    run_id: str | None = None,
    fail_on_issues: bool = False,
) -> int:
    ensure_dirs(input_dir, processed_dir, output_dir, config_dir, log_dir)
    log_path = setup_logging(log_dir, run_id=run_id)
    logging.info("Iniciando pipeline MVP IRONCORE")
    logging.info("Dirs: input=%s processed=%s output=%s config=%s", input_dir, processed_dir, output_dir, config_dir)

    aliases, required_fields = load_mappings(config_dir)

    source_files = [Path(p) for p in glob.glob(str(input_dir / "*"))]
    csv_files = [p for p in source_files if p.suffix.lower() == ".csv"]
    xlsx_files = [p for p in source_files if p.suffix.lower() in {".xlsx", ".xlsm", ".xls"}]

    if not csv_files and not xlsx_files:
        logging.warning("Nenhum arquivo de entrada em %s", input_dir)
        return 2

    rows: List[Dict[str, Any]] = []
    issues: List[Dict[str, Any]] = []

    for f in csv_files:
        logging.info("Lendo CSV: %s", f.name)
        loaded_rows, load_issues = load_csv(f, aliases, required_fields)
        rows.extend(loaded_rows)
        issues.extend(load_issues)

    for f in xlsx_files:
        logging.info("Lendo XLSX: %s", f.name)
        loaded_rows, load_issues = load_xlsx(f, aliases, required_fields)
        rows.extend(loaded_rows)
        issues.extend(load_issues)

    valid_rows, validation_issues = validate_rows(rows, required_fields)
    issues.extend(validation_issues)

    facts = build_facts(valid_rows)
    risks = build_risks(facts, load_rules(config_dir))

    (processed_dir / "issues.json").write_text(json.dumps(issues, ensure_ascii=False, indent=2), encoding="utf-8")
    (processed_dir / "facts.jsonl").write_text("\n".join(json.dumps(f, ensure_ascii=False) for f in facts), encoding="utf-8")
    (processed_dir / "risk_register.json").write_text(json.dumps(risks, ensure_ascii=False, indent=2), encoding="utf-8")

    summary = {
        "processed": len(facts),
        "risks": len(risks),
        "critical_high": len([r for r in risks if r["level"] in {"Crítico", "Alto"}]),
        "issues": len(issues),
        "generated_at": dt.datetime.now().isoformat(),
        "log_file": log_path.name,
        "run_id": run_id,
    }

    report = {"summary": summary, "top_risks": risks[:max_risks], "issues": issues}
    (output_dir / "comite.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    (output_dir / "comite.md").write_text(render_markdown(summary, risks[:max_risks]), encoding="utf-8")

    logging.info("Pipeline concluído. Processados=%s Riscos=%s Issues=%s", summary["processed"], summary["risks"], summary["issues"])
    if fail_on_issues and issues:
        logging.error("Execução marcada como falha por issues de dados (--fail-on-issues).")
        return 3
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(description="IRONCORE MVP pipeline")
    parser.add_argument("--max-risks", type=int, default=10)
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_SOURCES)
    parser.add_argument("--processed-dir", type=Path, default=DEFAULT_PROCESSED)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUTS)
    parser.add_argument("--config-dir", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--log-dir", type=Path, default=DEFAULT_LOGS)
    parser.add_argument("--run-id", type=str, default=None)
    parser.add_argument("--fail-on-issues", action="store_true")
    args = parser.parse_args()

    raise SystemExit(
        run(
            max_risks=args.max_risks,
            input_dir=args.input_dir,
            processed_dir=args.processed_dir,
            output_dir=args.output_dir,
            config_dir=args.config_dir,
            log_dir=args.log_dir,
            run_id=args.run_id,
            fail_on_issues=args.fail_on_issues,
        )
    )


if __name__ == "__main__":
    main()
